import sys
import os
import json
import webbrowser
import google.generativeai as genai
from dotenv import load_dotenv
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5 import uic
from pathlib import Path
import PyPDF2
import docx
import openpyxl

# Document processing imports
import PyPDF2
import docx
import openpyxl
from pathlib import Path

# Load environment variables
load_dotenv()

class QuestionGenerator(QThread):
    """Thread til at generere spørgsmål i baggrunden"""
    question_ready = pyqtSignal(dict)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, topic, api_key, difficulty="let"):
        super().__init__()
        self.topic = topic
        self.api_key = api_key
        self.difficulty = difficulty
        self.model = None
        
    def run(self):
        try:
            # Konfigurer Gemini API
            if not self.api_key or self.api_key.strip() == '':
                self.error_occurred.emit("Gemini API key er ikke indtastet. Indtast venligst din API key.")
                return
                
            genai.configure(api_key=self.api_key)
            # Opdateret til den nye model navn
            self.model = genai.GenerativeModel('gemini-1.5-flash')
            
            # Prompt til at generere spørgsmål
            difficulty_instruction = ""
            if self.difficulty == "let":
                difficulty_instruction = "Spørgsmålet skal være let til moderat i sværhedsgrad. Fokuser på grundlæggende begreber og ikke for komplekse detaljer."
            else:
                difficulty_instruction = "Spørgsmålet skal være svært og udfordrende. Inkluder detaljerede og komplekse aspekter af emnet."
            
            if self.topic.startswith("Dokument indhold: "):
                # Hvis det er dokumentindhold, lav et mere specifikt prompt
                document_content = self.topic[18:]  # Fjern "Dokument indhold: " prefix
                prompt = f"""
                Baseret på følgende tekst, lav et multiple choice spørgsmål der tester forståelsen af indholdet:

                TEKST:
                {document_content[:3000]}...

                SVÆRHEDSGRAD: {difficulty_instruction}

                Formatet skal være JSON og se sådan ud:
                {{
                    "question": "Spørgsmålet her",
                    "options": {{
                        "A": "Svarmulighed A",
                        "B": "Svarmulighed B", 
                        "C": "Svarmulighed C",
                        "D": "Svarmulighed D"
                    }},
                    "correct_answer": "A",
                    "explanation": "Forklaring på hvorfor dette er det rigtige svar"
                }}

                Spørgsmålet skal være baseret direkte på informationen i teksten.
                Sørg for at svarmuligheder er relevante og at kun én er korrekt.
                Sørg for at svaret er på dansk.
                Returner kun JSON - ingen andre tekst eller formatering.
                """
            else:
                # Standard emne-prompt
                prompt = f"""
                Lav et multiple choice spørgsmål om emnet: {self.topic}
                
                SVÆRHEDSGRAD: {difficulty_instruction}
                
                Formatet skal være JSON og se sådan ud:
                {{
                    "question": "Spørgsmålet her",
                    "options": {{
                        "A": "Svarmulighed A",
                        "B": "Svarmulighed B", 
                        "C": "Svarmulighed C",
                        "D": "Svarmulighed D"
                    }},
                    "correct_answer": "A",
                    "explanation": "Forklaring på hvorfor dette er det rigtige svar"
                }}
                
                Sørg for at spørgsmålet matcher den angivne sværhedsgrad og at forklaringen er detaljeret.
                Sørg for at svaret er på dansk.
                Returner kun JSON - ingen andre tekst eller formatering.
                """
            
            response = self.model.generate_content(prompt)
            
            # Parse JSON response
            response_text = response.text.strip()
            
            # Fjern eventuelle markdown kodeblokke
            if response_text.startswith('```json'):
                response_text = response_text[7:-3].strip()
            elif response_text.startswith('```'):
                response_text = response_text[3:-3].strip()
                
            question_data = json.loads(response_text)
            
            # Validér data struktur
            required_keys = ['question', 'options', 'correct_answer', 'explanation']
            if not all(key in question_data for key in required_keys):
                raise ValueError("Ugyldig JSON struktur fra API")
                
            self.question_ready.emit(question_data)
            
        except json.JSONDecodeError as e:
            self.error_occurred.emit(f"Fejl ved parsing af AI svar: {str(e)}")
        except Exception as e:
            self.error_occurred.emit(f"Fejl ved generering af spørgsmål: {str(e)}")

class MainWindow(QMainWindow):
    """Hovedvindue til at vælge emne"""
    
    def __init__(self):
        super().__init__()
        self.load_ui()
        self.setup_connections()
        
    def load_ui(self):
        """Load UI fra .ui fil"""
        try:
            uic.loadUi('main_window.ui', self)
        except Exception as e:
            QMessageBox.critical(self, "UI Fejl", f"Kunne ikke loade UI fil: {str(e)}")
            sys.exit(1)
        
    def setup_connections(self):
        """Setup signal connections"""
        self.start_quiz_btn.clicked.connect(self.start_quiz)
        self.get_api_key_btn.clicked.connect(self.open_api_key_page)
        self.upload_document_btn.clicked.connect(self.upload_document)
        
    def open_api_key_page(self):
        """Åbn Google AI Studio for at få API key"""
        try:
            webbrowser.open("https://makersuite.google.com/app/apikey")
            
            # Vis instruktioner
            QMessageBox.information(
                self, 
                "Få API Key", 
                "Trin for at få din API key:\n\n"
                "1. Log ind med din Google konto\n"
                "2. Klik 'Create API key'\n"
                "3. Vælg dit projekt (eller opret nyt)\n"
                "4. Kopier API key'en\n"
                "5. Indsæt den i API Key feltet\n\n"
                "Siden åbner nu i din browser..."
            )
        except Exception as e:
            QMessageBox.warning(self, "Fejl", f"Kunne ikke åbne browser: {str(e)}")
        
    def start_quiz(self):
        """Start quiz med valgt emne eller uploadet dokument"""
        api_key = self.api_key_input.text().strip()
        topic = self.topic_input.text().strip()
        document_path = self.document_path_input.text().strip()
        
        # Tjek sværhedsgrad
        difficulty = "let" if self.easy_difficulty_btn.isChecked() else "svær"
        
        if not api_key:
            QMessageBox.warning(self, "Manglende API Key", 
                              "Indtast venligst din Gemini API key før du starter.")
            return
        
        # Tjek at enten emne eller dokument er angivet
        if not topic and not document_path:
            QMessageBox.warning(self, "Manglende Input", 
                              "Indtast venligst enten et emne eller upload et dokument.")
            return
        
        if topic and document_path:
            QMessageBox.warning(self, "For meget Input", 
                              "Vælg enten at indtaste et emne ELLER uploade et dokument - ikke begge dele.")
            return
        
        try:
            if document_path:
                # Udtræk indhold fra dokument
                document_content = self.extract_document_content(document_path)
                if not document_content.strip():
                    QMessageBox.warning(self, "Tomt Dokument", 
                                      "Det valgte dokument indeholder ingen tekst.")
                    return
                
                # Brug dokumentindhold som "emne"
                quiz_content = f"Dokument indhold: {document_content}"
                display_name = Path(document_path).name
            else:
                # Brug det indtastede emne
                quiz_content = topic
                display_name = topic
            
            # Start quiz vindue med sværhedsgrad
            self.quiz_window = QuizWindow(quiz_content, api_key, display_name, difficulty)
            self.quiz_window.show()
            self.hide()
            
        except Exception as e:
            QMessageBox.critical(self, "Fejl", f"Fejl ved behandling af dokument: {str(e)}")

    def extract_text_from_pdf(self, file_path):
        """Udtræk tekst fra PDF fil"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text.strip()
        except Exception as e:
            raise Exception(f"Fejl ved læsning af PDF: {str(e)}")
    
    def extract_text_from_word(self, file_path):
        """Udtræk tekst fra Word dokument"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            raise Exception(f"Fejl ved læsning af Word dokument: {str(e)}")
    
    def extract_text_from_excel(self, file_path):
        """Udtræk tekst fra Excel fil"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Ark: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        text += " | ".join(row_text) + "\n"
                text += "\n"
            
            return text.strip()
        except Exception as e:
            raise Exception(f"Fejl ved læsning af Excel fil: {str(e)}")
    
    def extract_document_content(self, file_path):
        """Udtræk indhold fra dokument baseret på filtype"""
        file_extension = Path(file_path).suffix.lower()
        
        if file_extension == '.pdf':
            return self.extract_text_from_pdf(file_path)
        elif file_extension in ['.docx', '.doc']:
            return self.extract_text_from_word(file_path)
        elif file_extension in ['.xlsx', '.xls']:
            return self.extract_text_from_excel(file_path)
        else:
            # For almindelige tekstfiler
            if file_extension == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read().strip()
            raise Exception(f"Ikke understøttet filtype: {file_extension}")

    def upload_document(self):
        """Håndter dokumentupload"""
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self,
            "Vælg dokument",
            "",
            "Dokumenter (*.pdf *.docx *.doc *.xlsx *.xls *.txt);;PDF filer (*.pdf);;Word dokumenter (*.docx *.doc);;Excel filer (*.xlsx *.xls);;Tekst filer (*.txt)"
        )
        
        if file_path:
            # Vis filsti i input feltet
            self.document_path_input.setText(file_path)
            
            # Ryd emne input hvis der er valgt et dokument
            self.topic_input.clear()

class QuizWindow(QMainWindow):
    """Quiz vindue til at vise spørgsmål"""
    
    def __init__(self, topic, api_key, display_name=None, difficulty="let"):
        super().__init__()
        self.topic = topic
        self.api_key = api_key
        self.display_name = display_name or topic
        self.difficulty = difficulty
        self.current_question = None
        self.selected_answer = None
        self.question_generator = None
        
        self.load_ui()
        self.setup_connections()
        # Vis både emne og sværhedsgrad
        difficulty_emoji = "😊" if difficulty == "let" else "🔥"
        self.topic_label.setText(f"Emne: {self.display_name} | Sværhedsgrad: {difficulty_emoji} {difficulty.title()}")
        self.generate_new_question()
        
    def load_ui(self):
        """Load UI fra .ui fil"""
        try:
            uic.loadUi('quiz_window.ui', self)
        except Exception as e:
            QMessageBox.critical(self, "UI Fejl", f"Kunne ikke loade UI fil: {str(e)}")
            sys.exit(1)
    
    def setup_connections(self):
        """Setup signal connections"""
        self.next_btn.clicked.connect(self.generate_new_question)
        
        # Radio button connections - automatisk tjek svar når valgt
        self.option_a_btn.clicked.connect(lambda: self.select_and_check_answer('A'))
        self.option_b_btn.clicked.connect(lambda: self.select_and_check_answer('B'))
        self.option_c_btn.clicked.connect(lambda: self.select_and_check_answer('C'))
        self.option_d_btn.clicked.connect(lambda: self.select_and_check_answer('D'))
        
    def select_and_check_answer(self, answer):
        """Vælg svar og tjek automatisk"""
        # Tjek at der er et spørgsmål at besvare
        if not self.current_question:
            return
            
        # Hvis der allerede er givet et svar, ignorer klik
        if self.selected_answer:
            return
            
        self.selected_answer = answer
        self.check_answer()
        
    def generate_new_question(self):
        """Generer nyt spørgsmål"""
        self.reset_ui()
        self.question_label.setText("Genererer spørgsmål...")
        
        # Start question generator thread med sværhedsgrad
        self.question_generator = QuestionGenerator(self.topic, self.api_key, self.difficulty)
        self.question_generator.question_ready.connect(self.display_question)
        self.question_generator.error_occurred.connect(self.handle_error)
        self.question_generator.start()
        
    def display_question(self, question_data):
        """Vis spørgsmål i UI"""
        self.current_question = question_data
        
        # Set question text
        self.question_label.setText(question_data['question'])
        
        # Set options og aktiver dem
        self.option_a_btn.setText(f"A) {question_data['options']['A']}")
        self.option_b_btn.setText(f"B) {question_data['options']['B']}")
        self.option_c_btn.setText(f"C) {question_data['options']['C']}")
        self.option_d_btn.setText(f"D) {question_data['options']['D']}")
        
        # Aktiver svar-knapperne
        self.option_a_btn.setEnabled(True)
        self.option_b_btn.setEnabled(True)
        self.option_c_btn.setEnabled(True)
        self.option_d_btn.setEnabled(True)
        
    def check_answer(self):
        """Check if selected answer is correct"""
        if not self.selected_answer:
            return
            
        correct_answer = self.current_question['correct_answer']
        is_correct = self.selected_answer == correct_answer
        
        # Deaktiver alle svar-knapper efter svar er givet
        self.option_a_btn.setEnabled(False)
        self.option_b_btn.setEnabled(False)
        self.option_c_btn.setEnabled(False)
        self.option_d_btn.setEnabled(False)
        
        # Show result
        explanation_text = f"Forklaring: {self.current_question['explanation']}"
        if is_correct:
            self.result_label.setText(f"✅ RIGTIGT!\n\n{explanation_text}")
            self.result_label.setStyleSheet("color: green; font-weight: bold; font-size: 14px; background-color: white; border: 1px solid #ddd; border-radius: 5px; padding: 10px;")
        else:
            self.result_label.setText(f"❌ FORKERT! Det rigtige svar er {correct_answer}\n\n{explanation_text}")
            self.result_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px; background-color: white; border: 1px solid #ddd; border-radius: 5px; padding: 10px;")
        
        # Enable next button
        self.next_btn.setEnabled(True)
        
    def reset_ui(self):
        """Reset UI for new question"""
        self.selected_answer = None
        self.result_label.setText("")
        
        # Clear radio buttons
        self.option_a_btn.setAutoExclusive(False)
        self.option_b_btn.setAutoExclusive(False)
        self.option_c_btn.setAutoExclusive(False)
        self.option_d_btn.setAutoExclusive(False)
        
        self.option_a_btn.setChecked(False)
        self.option_b_btn.setChecked(False)
        self.option_c_btn.setChecked(False)
        self.option_d_btn.setChecked(False)
        
        self.option_a_btn.setAutoExclusive(True)
        self.option_b_btn.setAutoExclusive(True)
        self.option_c_btn.setAutoExclusive(True)
        self.option_d_btn.setAutoExclusive(True)
        
        # Reset buttons - fjernet submit_btn
        self.next_btn.setEnabled(False)
        
    def handle_error(self, error_message):
        """Handle errors from question generator"""
        self.question_label.setText("Fejl ved indlæsning af spørgsmål")
        QMessageBox.critical(self, "Fejl ved generering af spørgsmål", error_message)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
